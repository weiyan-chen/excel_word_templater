import logging
from pathlib import Path

import openpyxl
from docxtpl import DocxTemplate

import logging_config

logger = logging.getLogger(__name__)


class ExcelWordTemplater:
    """
    A class to create Word documents from Excel(.xlsx) data and Word(.docx) templates.
    """

    def __init__(
        self,
        excel_path: str,
        template_column: str,
        output_column: str | None = None,
        default_output_name: str | None = None,
        data_folder: str | None = None,
        template_folder: str | None = None,
        output_folder: str | None = None,
    ) -> None:
        """
        Initializes the templater, loads Excel(.xlsx) data, and sets up folders.
        """
        self.excel_path: str = excel_path
        logger.info(f"{self.excel_path=}")
        self.data: list[dict[str, str]] = self.read_excel()

        self.template_column: str = template_column
        logger.info(f"{self.template_column=}")
        self.check_template_column()

        self.output_column: str | None = output_column
        logger.info(f"{self.output_column=}")

        self.default_output_name: str = default_output_name or "output"
        logger.info(f"{self.default_output_name=}")
        self.default_output_name_index: int = 1

        self.data_folder: str = data_folder or "./data"
        logger.info(f"{self.data_folder=}")

        self.template_folder: str = (
            f"{self.data_folder}/{template_folder}"
            if template_folder
            else f"{self.data_folder}/templates"
        )
        logger.info(f"{self.template_folder=}")

        self.output_folder: str = (
            f"{self.data_folder}/{output_folder}"
            if output_folder
            else f"{self.data_folder}/output"
        )
        logger.info(f"{self.output_folder=}")

        self.create_folders()

    def check_template_column(self) -> None:
        """
        Ensures the template_column exists in the data.
        """
        logger.info("Checking template column...")

        if not self.data or self.template_column not in self.data[0]:
            msg = f"template_column '{self.template_column}' not found in data headers."
            logger.error(msg)
            raise ValueError(msg)

        logger.info("Template column check passed.")

    def create_folders(self) -> None:
        """
        Creates necessary folders if they do not exist.
        """
        logger.info("Creating necessary folders...")

        for folder in [self.data_folder, self.template_folder, self.output_folder]:
            Path(folder).mkdir(exist_ok=True)

        logger.info("Folders created.")

    def read_excel(self) -> list[dict]:
        """
        Reads the Excel(.xlsx) file and returns a list of dictionaries where each dictionary represents a row.
        """
        logger.info("Reading Excel file...")

        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            raise e

        ws = wb.active
        if ws is None or ws.max_row <= 1:
            msg = "Worksheet is empty or has no data rows."
            logger.warning(msg)
            raise ValueError(msg)

        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[0])
        data_rows = rows[1:]
        # Replace None values with an empty string to avoid issues later.
        data: list[dict] = [
            {k: (v if v is not None else "") for k, v in zip(headers, row)}
            for row in data_rows
        ]

        logger.info("Excel file read.")

        return data

    def render_template(self, data: dict[str, str]) -> DocxTemplate:
        """
        Renders a Word(.docx) template using the provided data.
        """
        logger.info("Rendering template...")

        template_name: str = data[self.template_column]
        template_path = Path(self.template_folder) / f"{template_name}.docx"
        if not template_name or not template_path.exists():
            msg = f"Template file path '{template_path}' is invalid or does not exist."
            logger.error(msg)
            raise FileNotFoundError(msg)

        try:
            doc = DocxTemplate(template_path)
            doc.render(data)
            logger.info("Template rendered.")
            return doc
        except Exception as e:
            logger.error(f"Error rendering template: {e}")
            raise e

    def save_docx(self, data: dict[str, str], doc: DocxTemplate) -> str:
        """
        Saves the rendered Word(.docx) file using either a specified or default output name.
        """
        logger.info("Saving document...")

        if (
            self.output_column
            and self.output_column in data
            and data[self.output_column] != ""
        ):
            output_name: str = data[self.output_column]
            output_path = Path(self.output_folder) / f"{output_name}.docx"
            index = 1
            while output_path.exists():
                output_path = Path(self.output_folder) / f"{output_name}_{index}.docx"
                index += 1
        else:
            output_path = (
                Path(self.output_folder)
                / f"{self.default_output_name}_{self.default_output_name_index}.docx"
            )
            self.default_output_name_index += 1

        try:
            doc.save(output_path)
            logger.info(f"Document saved as '{output_path}'")
            return str(output_path)
        except Exception as e:
            logger.error(f"Failed to save document: {e}")
            raise e

    def run(self) -> list[str]:
        """
        Processes each row in the Excel data to render and save Word(.docx) files.
        """
        output_paths: list[str] = []
        for index, row in enumerate(self.data):
            try:
                logger.info(f"Processing row {index}...")
                doc: DocxTemplate = self.render_template(row)
                output_path: str = self.save_docx(row, doc)
                output_paths.append(output_path)
                logger.info(f"Row {index} processed.")
            except Exception as e:
                logger.error(f"Error processing row {index}: {e}")

        return output_paths


if __name__ == "__main__":
    logging_config.setup_logging()

    try:
        logger.info("Starting Excel-Word Templater...")

        output_paths: list[str] = ExcelWordTemplater(
            excel_path="./data/excel/data.xlsx",
            template_column="template",
            output_column="output",
        ).run()

        logger.info(f"{output_paths=}")
        logger.info("Excel-Word Templater completed.")

    except Exception:
        logger.error("Excel-Word Templater failed.")
